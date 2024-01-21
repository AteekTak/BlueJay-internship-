# who has worked more than 14 hrs in a single shift

import openpyxl
from datetime import timedelta, datetime

# loading excel file
wb = openpyxl.load_workbook("data.xlsx")
sh = wb.active

# max rows and columns of the Excel file
max_row = sh.max_row
max_col = sh.max_column


def b():
    index = 2

    while index <= max_row:
        try:
            total_time = timedelta(hours=0, minutes=0)
            date_format = '%m/%d/%Y %I:%M %p'
            time_format = '%H:%M'
            user_id = sh.cell(index, 1).value
            user_name = sh.cell(index, 8).value
            user_time_card = sh.cell(index, 5).value
            user_date = sh.cell(index, 3).value

            if isinstance(user_date, str):
                user_date = datetime.strptime(user_date, date_format)

            if isinstance(user_time_card, str):
                user_time_card = datetime.strptime(user_time_card, time_format)

            user_date = user_date.date()
            user_time_card = user_time_card.time()

            for i in range(index+1, max_row + 1):
                curr = sh.cell(i, 1).value
                curr_time_card = sh.cell(index, 5).value
                curr_date = sh.cell(i, 3).value

                if isinstance(curr_time_card, str):
                    curr_time_card = datetime.strptime(curr_time_card, time_format)

                if isinstance(curr_date, str):
                    curr_date = datetime.strptime(curr_date, date_format)

                curr_date = curr_date.date()
                curr_time_card = curr_time_card.time()

                if user_id == curr and curr_date == user_date:
                    total_time = (timedelta(hours=user_time_card.hour, minutes=user_time_card.minute) +
                                  timedelta(hours=curr_time_card.hour, minutes=curr_time_card.minute))

                else:
                    index = i + 1
                    break

            if timedelta(hours=14) < total_time:
                print(f"position_id: {user_id}, name: {user_name}")

        except Exception as e:
            # print(e)
            # print(f"Error in date format of {sh.cell(index, 1).value} , index: {index}")
            index += 1


b()
