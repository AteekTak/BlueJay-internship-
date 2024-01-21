# who worked for 7 consecutive days

import openpyxl
from datetime import timedelta, datetime

# loading excel file
wb = openpyxl.load_workbook("data.xlsx")
sh = wb.active

# max rows and columns of the Excel file
max_row = sh.max_row
max_col = sh.max_column


def a():
    index = 2

    while index <= max_row:
        try:
            days = 1
            date_format = "%m/%d/%Y %I:%M %p"
            user_id = sh.cell(index, 1).value
            user_name = sh.cell(index, 8).value
            user_date = sh.cell(index, 3).value

            if isinstance(user_date, str):
                user_date = datetime.strptime(user_date, date_format).date()

            else:
                user_date = user_date.date()

            for i in range(index, max_row + 1):
                curr = sh.cell(i, 1).value
                curr_date = sh.cell(i, 3).value

                if isinstance(curr_date, str):
                    curr_date = datetime.strptime(curr_date, date_format).date()

                else:
                    curr_date = curr_date.date()

                if user_id == curr and curr_date == user_date + timedelta(days=1):
                    days += 1
                    user_date = curr_date

                elif user_id == curr and curr_date == user_date:
                    continue

                elif user_id != curr:
                    index = i + 1
                    break
                else:
                    index = i + 1
                    break

            if days >= 7:
                print(f"position_id: {user_id}, name: {user_name}")

        except Exception as e:
            # print(e)
            # print(f"Error in date format of {sh.cell(index, 1).value} , index: {index}")
            index += 1


a()
